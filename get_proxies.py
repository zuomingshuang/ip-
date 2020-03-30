import openpyxl

def get_proxiex_list():
    wb=openpyxl.load_workbook('https_proxies-曾.xlsx')
    sht=wb.get_active_sheet()
    max_row=sht.max_row
    proxies_list=[]
    for r in range(2,max_row+1):
        proxies={'https':'https://'+sht.cell(r,3).value}
        proxies_list.append(proxies)
    return proxies_list
